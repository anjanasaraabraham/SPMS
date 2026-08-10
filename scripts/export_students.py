"""Export all seeded students to an XLSX file."""
import asyncio
import os
from pathlib import Path
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

ROOT = Path(__file__).resolve().parent.parent / "backend"
load_dotenv(ROOT / ".env")

MONGO_URL = os.environ["MONGO_URL"]
DB_NAME = os.environ["DB_NAME"]

OUT_PATH = "/app/SPMS_Students_Dataset.xlsx"


async def main():
    client = AsyncIOMotorClient(MONGO_URL)
    db = client[DB_NAME]
    students = await db.users.find({"role": "student"}, {"_id": 0, "password_hash": 0}).to_list(1000)
    # sort by roll number
    students.sort(key=lambda s: s.get("roll_number", ""))
    client.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Students"

    headers = ["S.No", "Roll Number", "Name", "Email", "Department", "Role", "Created At"]
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", fgColor="D34449")
    center = Alignment(horizontal="center", vertical="center")
    border_side = Side(style="thin", color="E5E7EB")
    thin = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    ws.append(headers)
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = center
        c.border = thin

    for idx, s in enumerate(students, start=1):
        ws.append([
            idx,
            s.get("roll_number", ""),
            s.get("name", ""),
            s.get("email", ""),
            s.get("department", "PGDM"),
            s.get("role", "student"),
            s.get("created_at", "")[:19].replace("T", " "),
        ])
        for col in range(1, len(headers) + 1):
            ws.cell(row=idx + 1, column=col).border = thin
            ws.cell(row=idx + 1, column=col).alignment = Alignment(vertical="center", horizontal="left" if col in (3, 4, 7) else "center")

    # Column widths
    widths = [8, 15, 26, 32, 14, 12, 22]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    wb.save(OUT_PATH)
    print(f"Saved {len(students)} students to {OUT_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
